"""
ETL: LOG_INs 2.xlsx -> PostgreSQL (Neon)

Imports SIM card lookup data used by Fahrer reports:
- sheet "Contado"
- sheet "Vodafone&O2 SIM-Karten  Neu"
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import shutil
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path

from dotenv import load_dotenv
import openpyxl


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SIM_CARDS_PATH = Path(
    r"C:\Users\Aleksei Samosvat\Groo GmbH\Communication site - Documents\Groo Cargo Logistic\GC_IT\GC_Sim-Karten_LOG_IN\LOG_INs 2.xlsx"
)
CONTADO_SHEET = "Contado"
VODAFONE_SHEET = "Vodafone&O2 SIM-Karten  Neu"


def _lazy_import_psycopg():
    try:
        import psycopg  # type: ignore
    except Exception as exc:  # pragma: no cover - runtime dependency message
        raise RuntimeError(
            "psycopg is required. Install dependencies from requirements.txt "
            "and ensure DATABASE_URL is set."
        ) from exc
    return psycopg


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _norm_lkw(value: object) -> str:
    return "".join(str(value or "").strip().upper().split())


def _prepare_readable_xlsx(source_path: Path) -> tuple[Path, bool]:
    """
    Returns (path, is_temp_copy_created_by_this_run).
    """
    temp_dir = Path(tempfile.gettempdir())
    run_copy = temp_dir / f"SIM_LOG_INs_ETL_{int(time.time())}.xlsx"
    try:
        shutil.copy2(source_path, run_copy)
        return run_copy, True
    except Exception:
        candidates = sorted(glob.glob(str(temp_dir / "SIM_LOG_INs_ETL_*.xlsx")), reverse=True)
        for candidate in candidates:
            path = Path(candidate)
            if path.exists() and path.stat().st_size > 0:
                return path, False
    raise PermissionError(
        "Could not create temp copy from SIM_CARDS_FILE_PATH and no fallback copy was found in %TEMP%."
    )


@dataclass
class ContadoRow:
    lkw_number: str
    name: str
    password: str
    source_row: int


@dataclass
class VodafoneRow:
    lkw_number: str
    pin: str
    puk: str
    source_row: int


def extract_contado_rows(wb) -> list[ContadoRow]:
    ws = wb[CONTADO_SHEET]
    by_lkw: dict[str, ContadoRow] = {}
    for row_idx in range(2, ws.max_row + 1):
        lkw_number = _clean_text(ws.cell(row_idx, 2).value)
        name = _clean_text(ws.cell(row_idx, 3).value)
        password = _clean_text(ws.cell(row_idx, 5).value)
        if not lkw_number:
            continue
        key = _norm_lkw(lkw_number)
        if not key:
            continue
        by_lkw[key] = ContadoRow(
            lkw_number=lkw_number,
            name=name or "",
            password=password or "",
            source_row=row_idx,
        )
    return list(by_lkw.values())


def extract_vodafone_rows(wb) -> list[VodafoneRow]:
    ws = wb[VODAFONE_SHEET]
    by_lkw: dict[str, VodafoneRow] = {}
    for row_idx in range(2, ws.max_row + 1):
        lkw_number = _clean_text(ws.cell(row_idx, 5).value)
        pin = _clean_text(ws.cell(row_idx, 9).value)
        puk = _clean_text(ws.cell(row_idx, 10).value)
        if not lkw_number:
            continue
        key = _norm_lkw(lkw_number)
        if not key:
            continue
        by_lkw[key] = VodafoneRow(
            lkw_number=lkw_number,
            pin=pin or "",
            puk=puk or "",
            source_row=row_idx,
        )
    return list(by_lkw.values())


def _ensure_contado_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_sim_contado (
            lkw_number TEXT PRIMARY KEY,
            sim_name TEXT NOT NULL DEFAULT '',
            password TEXT NOT NULL DEFAULT '',
            source_row INTEGER NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_sim_contado_lkw
            ON report_sim_contado (lkw_number)
        """
    )


def _ensure_vodafone_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_sim_vodafone (
            lkw_number TEXT PRIMARY KEY,
            pin TEXT NOT NULL DEFAULT '',
            puk TEXT NOT NULL DEFAULT '',
            source_row INTEGER NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_sim_vodafone_lkw
            ON report_sim_vodafone (lkw_number)
        """
    )


def run_etl(database_url: str, source_path: Path) -> dict[str, int]:
    psycopg = _lazy_import_psycopg()
    created_copy = False
    readable_path = source_path
    log_id = None

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO etl_log (source_name, status, details)
                VALUES (%s, 'running', %s::jsonb)
                RETURNING id
                """,
                ("xlsx_sim_cards", json.dumps({"source_path": str(source_path)}, ensure_ascii=False)),
            )
            log_id = int(cur.fetchone()[0])
        conn.commit()

        try:
            readable_path, created_copy = _prepare_readable_xlsx(source_path)
            wb = openpyxl.load_workbook(readable_path, read_only=True, data_only=True)
            contado_rows = extract_contado_rows(wb)
            vodafone_rows = extract_vodafone_rows(wb)
            wb.close()

            with conn.cursor() as cur:
                _ensure_contado_table(cur)
                _ensure_vodafone_table(cur)
                cur.execute(
                    """
                    CREATE TEMP TABLE tmp_report_sim_contado
                    (LIKE report_sim_contado INCLUDING DEFAULTS)
                    ON COMMIT DROP
                    """
                )
                cur.execute(
                    """
                    CREATE TEMP TABLE tmp_report_sim_vodafone
                    (LIKE report_sim_vodafone INCLUDING DEFAULTS)
                    ON COMMIT DROP
                    """
                )

                for row in contado_rows:
                    cur.execute(
                        """
                        INSERT INTO tmp_report_sim_contado (
                            lkw_number, sim_name, password, source_row, raw_payload, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            row.lkw_number,
                            row.name,
                            row.password,
                            row.source_row,
                            json.dumps(
                                {
                                    "sheet": CONTADO_SHEET,
                                    "LKW": row.lkw_number,
                                    "Name": row.name,
                                    "Password": row.password,
                                },
                                ensure_ascii=False,
                            ),
                        ),
                    )

                for row in vodafone_rows:
                    cur.execute(
                        """
                        INSERT INTO tmp_report_sim_vodafone (
                            lkw_number, pin, puk, source_row, raw_payload, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            row.lkw_number,
                            row.pin,
                            row.puk,
                            row.source_row,
                            json.dumps(
                                {
                                    "sheet": VODAFONE_SHEET,
                                    "LKW Kennzeichen": row.lkw_number,
                                    "PIN": row.pin,
                                    "PUK": row.puk,
                                },
                                ensure_ascii=False,
                            ),
                        ),
                    )

                cur.execute("DELETE FROM report_sim_contado")
                cur.execute("INSERT INTO report_sim_contado SELECT * FROM tmp_report_sim_contado")
                cur.execute("DELETE FROM report_sim_vodafone")
                cur.execute("INSERT INTO report_sim_vodafone SELECT * FROM tmp_report_sim_vodafone")

                cur.execute(
                    """
                    UPDATE etl_log
                    SET status = 'success',
                        finished_at = NOW(),
                        rows_inserted = %s,
                        details = %s::jsonb
                    WHERE id = %s
                    """,
                    (
                        len(contado_rows) + len(vodafone_rows),
                        json.dumps(
                            {
                                "workbook_used": str(readable_path),
                                "contado_rows": len(contado_rows),
                                "vodafone_rows": len(vodafone_rows),
                            },
                            ensure_ascii=False,
                        ),
                        log_id,
                    ),
                )
            conn.commit()
            return {
                "contado_rows": len(contado_rows),
                "vodafone_rows": len(vodafone_rows),
            }
        except Exception as exc:
            conn.rollback()
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE etl_log
                    SET status = 'failed',
                        finished_at = NOW(),
                        error_message = %s,
                        details = %s::jsonb
                    WHERE id = %s
                    """,
                    (
                        str(exc),
                        json.dumps(
                            {
                                "source_path": str(source_path),
                                "workbook_used": str(readable_path),
                            },
                            ensure_ascii=False,
                        ),
                        log_id,
                    ),
                )
            conn.commit()
            raise
        finally:
            if created_copy and readable_path.exists():
                try:
                    readable_path.unlink()
                except Exception:
                    pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Import SIM card workbook into PostgreSQL")
    parser.add_argument("--database-url", default="", help="Override DATABASE_URL from env")
    parser.add_argument("--source", default="", help="Override SIM_CARDS_FILE_PATH / default path")
    args = parser.parse_args()

    load_dotenv(BASE_DIR / ".env", override=True)
    database_url = (args.database_url or os.getenv("DATABASE_URL", "")).strip()
    if not database_url:
        raise RuntimeError("DATABASE_URL is empty. Set it in .env or pass --database-url.")

    source_path = Path(
        os.path.expandvars((args.source or os.getenv("SIM_CARDS_FILE_PATH") or str(DEFAULT_SIM_CARDS_PATH)).strip())
    )
    if not source_path.exists():
        raise FileNotFoundError(f"SIM cards workbook not found: {source_path}")

    result = run_etl(database_url, source_path)
    print(
        "ETL success: "
        f"contado_rows={result['contado_rows']} "
        f"vodafone_rows={result['vodafone_rows']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
