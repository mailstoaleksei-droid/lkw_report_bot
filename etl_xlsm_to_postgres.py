"""
ETL: LKW_Fahrer_Data.xlsm -> PostgreSQL (Neon)

Phase 1.3 MVP:
- reads master data from sheets "LKW" and "Fahrer"
- reads monthly revenue data from sheet "Bericht_Dispo"
- reads monthly bonus dynamics from sheet "BonusDynamik"
- reads monthly diesel data from sheet "Diesel"
- upserts companies, trucks, drivers
- writes run metadata to etl_log
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import re
import shutil
import tempfile
import time
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

from dotenv import load_dotenv
import openpyxl


def _lazy_import_psycopg():
    try:
        import psycopg  # type: ignore
    except Exception as exc:  # pragma: no cover - runtime dependency message
        raise RuntimeError(
            "psycopg is required. Install dependencies from requirements.txt "
            "and ensure DATABASE_URL is set."
        ) from exc
    return psycopg


REQUIRED_TRUCK_KEYS = ("lkwid", "lkwnummer")
REQUIRED_DRIVER_KEYS = ("fahrerid", "fahrername")
BERICHT_DISPO_SHEET = "Bericht_Dispo"
BONUS_DYNAMIK_SHEET = "BonusDynamik"
DIESEL_SHEET = "Diesel"
YF_FAHRER_SHEET = "YF_Fahrer"
YF_SHEET = "YF"
REPAIR_SHEET = "Repair"
STAACK_SHEET = "Staack"
SHELL_SHEET = "Shell"
CARLO_SHEET = "Carlo"
CONTADO_SHEET = "Contado"
REPAIR_TRUCK_RENAMES = {
    "DE-FN186": "GR-OO2103",
    "DE-FN401": "GR-OO2104",
    "DE-FN179": "GR-OO2205",
    "DE-FN400": "GR-OO2206",
}
MONTHS_DE = {
    "januar": (1, "Januar"),
    "jan": (1, "Januar"),
    "january": (1, "Januar"),
    "februar": (2, "Februar"),
    "feb": (2, "Februar"),
    "february": (2, "Februar"),
    "maerz": (3, "Maerz"),
    "marz": (3, "Maerz"),
    "march": (3, "Maerz"),
    "april": (4, "April"),
    "apr": (4, "April"),
    "mai": (5, "Mai"),
    "may": (5, "Mai"),
    "juni": (6, "Juni"),
    "jun": (6, "Juni"),
    "june": (6, "Juni"),
    "juli": (7, "Juli"),
    "jul": (7, "Juli"),
    "july": (7, "Juli"),
    "august": (8, "August"),
    "aug": (8, "August"),
    "september": (9, "September"),
    "sep": (9, "September"),
    "sept": (9, "September"),
    "oktober": (10, "Oktober"),
    "okt": (10, "Oktober"),
    "october": (10, "Oktober"),
    "oct": (10, "Oktober"),
    "november": (11, "November"),
    "nov": (11, "November"),
    "dezember": (12, "Dezember"),
    "dez": (12, "Dezember"),
    "december": (12, "Dezember"),
    "dec": (12, "Dezember"),
}
MONTH_NAMES_DE = [
    "Januar",
    "Februar",
    "Maerz",
    "April",
    "Mai",
    "Juni",
    "Juli",
    "August",
    "September",
    "Oktober",
    "November",
    "Dezember",
]


def _norm(value: object) -> str:
    s = str(value or "").strip().lower()
    repl = (
        ("ä", "a"),
        ("ö", "o"),
        ("ü", "u"),
        ("ß", "ss"),
        ("\n", " "),
    )
    for a, b in repl:
        s = s.replace(a, b)
    return re.sub(r"[^a-z0-9]+", "", s)


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        d = value.date()
    elif isinstance(value, date):
        d = value
    else:
        s = str(value).strip()
        if not s or s == "00:00:00":
            return None
        if " " in s:
            s = s.split(" ", 1)[0]
        try:
            d = datetime.fromisoformat(s).date()
        except ValueError:
            d = None
            for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
            if d is None:
                return None
    if d.year < 1970 or d.year > 2100:
        return None
    return d


def _parse_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    raw = str(value).strip()
    if not raw:
        return Decimal("0.00")
    if raw.upper() in {"#REF!", "#N/A", "N/A"}:
        return Decimal("0.00")

    cleaned = raw.replace(" ", "").replace("\u00a0", "")
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    cleaned = re.sub(r"[^0-9.\-]+", "", cleaned)
    if cleaned in {"", "-", ".", "-."}:
        return Decimal("0.00")

    try:
        return Decimal(cleaned).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _normalize_repair_truck_number(value: object) -> str | None:
    raw = _clean_text(value)
    if not raw:
        return None
    compact = re.sub(r"\s+", " ", raw).strip().upper()
    if compact.startswith("EX "):
        compact = compact[3:].strip()
    return REPAIR_TRUCK_RENAMES.get(compact, compact)


def _normalize_lkw_number(value: object) -> str | None:
    raw = _clean_text(value)
    if not raw:
        return None
    compact = re.sub(r"\s+", " ", raw.strip().strip("'")).strip().upper()
    if compact.startswith("EX "):
        compact = compact[3:].strip()
    return REPAIR_TRUCK_RENAMES.get(compact, compact)


def _normalize_fuel_product(value: object) -> str | None:
    product = _clean_text(value)
    norm = _norm(product)
    if not norm:
        return None
    if "adblue" in norm:
        return "AdBlue"
    if "diesel" in norm:
        return "Diesel"
    return None


def _parse_month_token(value: object) -> tuple[int, str] | None:
    key = _norm(value)
    if not key:
        return None
    return MONTHS_DE.get(key)


def _find_metric_row(ws, aliases: Iterable[str], max_scan_rows: int = 120, max_scan_cols: int = 8) -> int | None:
    normalized_aliases = {_norm(a) for a in aliases if _norm(a)}
    if not normalized_aliases:
        return None
    for row_idx in range(1, max_scan_rows + 1):
        for col_idx in range(1, max_scan_cols + 1):
            token = _norm(ws.cell(row=row_idx, column=col_idx).value)
            if token in normalized_aliases:
                return row_idx
    return None


def _discover_month_columns(ws, preferred_row: int = 6, max_scan_rows: int = 30) -> dict[int, tuple[str, int]]:
    def parse_row(row_idx: int) -> dict[int, tuple[str, int]]:
        found: dict[int, tuple[str, int]] = {}
        for col_idx in range(1, ws.max_column + 1):
            parsed = _parse_month_token(ws.cell(row=row_idx, column=col_idx).value)
            if not parsed:
                continue
            month_idx, month_name = parsed
            if month_idx not in found:
                found[month_idx] = (month_name, col_idx)
        return found

    preferred = parse_row(preferred_row)
    if len(preferred) >= 10:
        return preferred

    best: dict[int, tuple[str, int]] = preferred
    for row_idx in range(1, max_scan_rows + 1):
        found = parse_row(row_idx)
        if len(found) > len(best):
            best = found
    return best


def _find_header_row(ws, required_keys: Iterable[str], max_scan_rows: int = 50) -> int:
    required = set(required_keys)
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > max_scan_rows:
            break
        normalized = {_norm(v) for v in row if _norm(v)}
        if required.issubset(normalized):
            return row_idx
    raise RuntimeError(f"Header row not found in sheet '{ws.title}' for keys: {sorted(required)}")


def _get_row_values(ws, row_idx: int) -> list[object]:
    return [cell for cell in next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))]


def _build_col_index(header_row: list[object]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, v in enumerate(header_row):
        k = _norm(v)
        if k and k not in index:
            index[k] = i
    return index


def _pick_col(index: dict[str, int], *aliases: str) -> int | None:
    for alias in aliases:
        k = _norm(alias)
        if k in index:
            return index[k]
    return None


@dataclass
class TruckRow:
    external_id: str
    plate_number: str | None
    truck_type: str | None
    company_name: str | None
    status: str | None
    status_since: date | None
    is_active: bool
    raw_payload: dict[str, str | None]


@dataclass
class DriverRow:
    external_id: str
    full_name: str
    company_name: str | None
    phone: str | None
    is_active: bool
    raw_payload: dict[str, str | None]


@dataclass
class FahrerWeekStatusRow:
    report_year: int
    iso_week: int
    week_start: date
    week_end: date
    fahrer_id: str
    fahrer_name: str
    company_name: str | None
    status_entlassen: str | None
    datum_entlassen: date | None
    week_code: str
    is_active_in_week: bool
    raw_payload: dict[str, object]


@dataclass
class EinnahmenMonthRow:
    month_index: int
    month_name: str
    nahverkehr: Decimal
    logistics: Decimal
    gesamt: Decimal
    raw_payload: dict[str, object]


@dataclass
class EinnahmenFirmRow:
    row_index: int
    firm_name: str
    january: Decimal
    february: Decimal
    march: Decimal
    april: Decimal
    may: Decimal
    june: Decimal
    july: Decimal
    august: Decimal
    september: Decimal
    october: Decimal
    november: Decimal
    december: Decimal
    total: Decimal
    raw_payload: dict[str, object]


@dataclass
class BonusDynamikRow:
    report_year: int
    report_month: int
    month_start: date
    fahrer_id: str
    fahrer_name: str
    days: int
    km: Decimal
    pct_km: Decimal
    ct: int
    pct_ct: Decimal
    bonus: Decimal
    penalty: Decimal
    final: Decimal
    raw_payload: dict[str, object]


@dataclass
class DieselMonthRow:
    report_year: int
    month_index: int
    month_name: str
    liter_staack: Decimal
    liter_shell: Decimal
    liter_dkv: Decimal
    liter_total: Decimal
    euro_staack: Decimal
    euro_shell: Decimal
    euro_dkv: Decimal
    euro_total: Decimal
    euro_per_liter_staack: Decimal
    euro_per_liter_shell: Decimal
    euro_per_liter_dkv: Decimal
    euro_per_liter_avg: Decimal
    raw_payload: dict[str, object]


@dataclass
class LkwFuelTransactionRow:
    source: str
    source_row: int
    report_year: int
    report_month: int
    iso_week: int
    transaction_date: date | None
    lkw_number: str
    product_name: str
    quantity_liters: Decimal
    total_net: Decimal
    driver_name: str | None
    raw_payload: dict[str, object]


@dataclass
class LkwRevenueRow:
    source: str
    source_row: int
    report_year: int
    report_month: int
    iso_week: int
    lkw_number: str
    revenue_amount: Decimal
    raw_payload: dict[str, object]


@dataclass
class YFFahrerMonthRow:
    month_index: int
    fahrer_name: str
    distanz_km: Decimal
    aktivitaet_total_minutes: int
    fahrzeit_total_minutes: int
    inaktivitaet_total_minutes: int
    raw_payload: dict[str, object]


@dataclass
class YFLkwDayRow:
    report_year: int
    month_index: int
    month_name: str
    iso_week: int
    lkw_nummer: str
    report_date: date
    source_row: int
    dayweek: str | None
    strecke_km: Decimal
    km_start: Decimal
    km_end: Decimal
    drivers_final: str | None
    raw_payload: dict[str, object]


@dataclass
class RepairRow:
    report_year: int
    report_month: int
    iso_week: int
    invoice_date: date | None
    truck_number: str
    original_truck_number: str | None
    repair_name: str | None
    total_price: Decimal
    invoice: str | None
    seller: str | None
    buyer: str | None
    kategorie: str | None
    source_row: int
    raw_payload: dict[str, object]


def _iter_sheet_rows(ws, header_row_idx: int):
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        yield list(row)


def extract_repairs(wb) -> list[RepairRow]:
    if REPAIR_SHEET not in wb.sheetnames:
        return []

    ws = wb[REPAIR_SHEET]
    header_row_idx = _find_header_row(ws, ("month", "week", "truck", "totalprice"))
    header = _get_row_values(ws, header_row_idx)
    index = _build_col_index(header)

    col_year = _pick_col(index, "Year", "Yaer")
    col_month = _pick_col(index, "Month")
    col_week = _pick_col(index, "Week")
    col_date = _pick_col(index, "Date Invoice", "Invoice Date", "Date")
    col_truck = _pick_col(index, "Truck")
    col_name = _pick_col(index, "Name")
    col_total = _pick_col(index, "Total Price", "TotalPrice", "Price", "Total")
    col_invoice = _pick_col(index, "Invoice")
    col_seller = _pick_col(index, "Seller")
    col_buyer = _pick_col(index, "Buyer", "Byuer")
    col_kategorie = _pick_col(index, "Kategorie", "Category")

    if col_truck is None or col_total is None:
        return []

    rows: list[RepairRow] = []
    for row_idx, row in enumerate(_iter_sheet_rows(ws, header_row_idx), start=header_row_idx + 1):
        if col_truck >= len(row):
            continue
        truck_raw = _clean_text(row[col_truck])
        truck_number = _normalize_repair_truck_number(truck_raw)
        if not truck_number:
            continue

        invoice_date = _parse_date(row[col_date]) if col_date is not None and col_date < len(row) else None
        report_year = (
            _parse_strict_positive_int(row[col_year])
            if col_year is not None and col_year < len(row)
            else None
        )
        report_month = (
            _parse_strict_positive_int(row[col_month])
            if col_month is not None and col_month < len(row)
            else None
        )
        iso_week = (
            _parse_strict_positive_int(row[col_week])
            if col_week is not None and col_week < len(row)
            else None
        )
        if invoice_date:
            report_year = report_year or invoice_date.year
            report_month = report_month or invoice_date.month
            iso_week = iso_week or int(invoice_date.isocalendar()[1])
        if not report_year or report_year < 2020 or report_year > 2100:
            continue
        if not report_month or report_month < 1 or report_month > 12:
            report_month = 1
        if not iso_week or iso_week < 1 or iso_week > 53:
            iso_week = 1

        payload = {
            str(header[i]).strip() if i < len(header) and header[i] is not None else f"col_{i+1}":
            row[i] if i < len(row) else None
            for i in range(len(header))
        }
        payload["normalized_truck_number"] = truck_number
        payload["source_row"] = row_idx

        rows.append(
            RepairRow(
                report_year=report_year,
                report_month=report_month,
                iso_week=iso_week,
                invoice_date=invoice_date,
                truck_number=truck_number,
                original_truck_number=truck_raw,
                repair_name=_clean_text(row[col_name]) if col_name is not None and col_name < len(row) else None,
                total_price=_parse_decimal(row[col_total]) if col_total < len(row) else Decimal("0.00"),
                invoice=_clean_text(row[col_invoice]) if col_invoice is not None and col_invoice < len(row) else None,
                seller=_clean_text(row[col_seller]) if col_seller is not None and col_seller < len(row) else None,
                buyer=_clean_text(row[col_buyer]) if col_buyer is not None and col_buyer < len(row) else None,
                kategorie=_clean_text(row[col_kategorie]) if col_kategorie is not None and col_kategorie < len(row) else None,
                source_row=row_idx,
                raw_payload=payload,
            )
        )

    return sorted(rows, key=lambda r: (r.report_year, r.report_month, r.iso_week, r.truck_number, r.source_row))


def extract_trucks(wb) -> list[TruckRow]:
    ws = wb["LKW"]
    header_row_idx = _find_header_row(ws, REQUIRED_TRUCK_KEYS)
    header = _get_row_values(ws, header_row_idx)
    index = _build_col_index(header)

    col_id = _pick_col(index, "LKW-ID", "ID")
    col_num = _pick_col(index, "LKW-Nummer", "Number")
    col_type = _pick_col(index, "LKW-Typ", "Type")
    col_company = _pick_col(index, "Firma", "Company")
    col_status = _pick_col(index, "Status")
    col_sale_date = _pick_col(index, "Datum verkauft", "Sale Date")

    rows: list[TruckRow] = []
    for row in _iter_sheet_rows(ws, header_row_idx):
        if col_id is None or col_id >= len(row):
            continue
        external_id = _clean_text(row[col_id])
        if not external_id or not external_id.upper().startswith("L"):
            continue

        status = _clean_text(row[col_status]) if col_status is not None and col_status < len(row) else None
        status_norm = _norm(status)
        is_active = status_norm not in {"verkauft", "ruckgabe", "rueckgabe", "sold", "inactive"}

        payload = {
            str(header[i]).strip() if i < len(header) and header[i] is not None else f"col_{i+1}":
            _clean_text(row[i]) if i < len(row) else None
            for i in range(len(header))
        }

        rows.append(
            TruckRow(
                external_id=external_id,
                plate_number=_clean_text(row[col_num]) if col_num is not None and col_num < len(row) else None,
                truck_type=_clean_text(row[col_type]) if col_type is not None and col_type < len(row) else None,
                company_name=_clean_text(row[col_company]) if col_company is not None and col_company < len(row) else None,
                status=status,
                status_since=_parse_date(row[col_sale_date]) if col_sale_date is not None and col_sale_date < len(row) else None,
                is_active=is_active,
                raw_payload=payload,
            )
        )
    return rows


def extract_drivers(wb) -> list[DriverRow]:
    ws = wb["Fahrer"]
    header_row_idx = _find_header_row(ws, REQUIRED_DRIVER_KEYS)
    header = _get_row_values(ws, header_row_idx)
    index = _build_col_index(header)

    col_id = _pick_col(index, "Fahrer-ID", "ID")
    col_name = _pick_col(index, "Fahrername", "Name")
    col_company = _pick_col(index, "Firma", "Company")
    col_phone = _pick_col(index, "Telefonnummer", "Phone")
    col_status = _pick_col(index, "Status", "Active/Fired")

    year_headers = _effective_header_values(ws, max(1, header_row_idx - 1))
    vacation_cols_by_year: dict[int, int] = {}
    sick_cols_by_year: dict[int, int] = {}
    for i, label in enumerate(header):
        report_year = _parse_strict_positive_int(year_headers[i]) if i < len(year_headers) else None
        if not report_year or report_year < 2020 or report_year > 2100:
            continue
        label_norm = _norm(label)
        if label_norm == _norm("Urlaub gesamt"):
            vacation_cols_by_year[report_year] = i
        elif label_norm == _norm("Krankheitstage"):
            sick_cols_by_year[report_year] = i
    totals_year = max([*vacation_cols_by_year.keys(), *sick_cols_by_year.keys()], default=None)
    col_vacation_total = vacation_cols_by_year.get(totals_year) if totals_year else None
    col_sick_days = sick_cols_by_year.get(totals_year) if totals_year else None

    rows: list[DriverRow] = []
    for row in _iter_sheet_rows(ws, header_row_idx):
        if col_id is None or col_name is None:
            continue
        if col_id >= len(row) or col_name >= len(row):
            continue
        external_id = _clean_text(row[col_id])
        full_name = _clean_text(row[col_name])
        if not external_id or not full_name or not external_id.upper().startswith("F"):
            continue

        status = _clean_text(row[col_status]) if col_status is not None and col_status < len(row) else None
        status_norm = _norm(status)
        is_active = status_norm not in {"fired", "entlassen", "inactive", "inaktiv"}

        payload = {
            str(header[i]).strip() if i < len(header) and header[i] is not None else f"col_{i+1}":
            _clean_text(row[i]) if i < len(row) else None
            for i in range(len(header))
        }
        if totals_year:
            payload[f"Urlaub gesamt {totals_year}"] = (
                _clean_text(row[col_vacation_total])
                if col_vacation_total is not None and col_vacation_total < len(row)
                else None
            )
            payload[f"Krankheitstage {totals_year}"] = (
                _clean_text(row[col_sick_days])
                if col_sick_days is not None and col_sick_days < len(row)
                else None
            )

        rows.append(
            DriverRow(
                external_id=external_id,
                full_name=full_name,
                company_name=_clean_text(row[col_company]) if col_company is not None and col_company < len(row) else None,
                phone=_clean_text(row[col_phone]) if col_phone is not None and col_phone < len(row) else None,
                is_active=is_active,
                raw_payload=payload,
            )
        )
    return rows


def _normalize_fahrer_week_code(value: object) -> str:
    raw = str(value or "").strip().upper()
    if raw == "К":
        return "K"
    if raw in {"U", "K"}:
        return raw
    return ""


def extract_fahrer_weekly_statuses(wb) -> list[FahrerWeekStatusRow]:
    ws = wb["Fahrer"]
    header_row_idx = _find_header_row(ws, REQUIRED_DRIVER_KEYS)
    header = _get_row_values(ws, header_row_idx)
    index = _build_col_index(header)

    year_headers = _effective_header_values(ws, header_row_idx)
    sub_headers = _get_row_values(ws, header_row_idx + 1)

    col_id = _pick_col(index, "Fahrer-ID", "ID")
    col_name = _pick_col(index, "Fahrername", "Name")
    col_company = _pick_col(index, "Firma", "Company")
    col_status = _pick_col(index, "Status", "Active/Fired")
    col_dismiss = _pick_col(index, "Datum entlassen", "Date")

    week_columns: list[tuple[int, int, int]] = []
    for col_idx in range(1, int(ws.max_column or 0) + 1):
        report_year = _parse_strict_positive_int(year_headers[col_idx - 1]) if col_idx - 1 < len(year_headers) else None
        iso_week = _parse_strict_positive_int(sub_headers[col_idx - 1]) if col_idx - 1 < len(sub_headers) else None
        if not report_year or report_year < 2020 or report_year > 2100 or not iso_week or iso_week < 1 or iso_week > 53:
            continue
        week_columns.append((col_idx - 1, report_year, iso_week))

    if col_id is None or col_name is None or not week_columns:
        return []

    rows: list[FahrerWeekStatusRow] = []
    for row_idx, row in enumerate(_iter_sheet_rows(ws, header_row_idx), start=header_row_idx + 1):
        if col_id >= len(row) or col_name >= len(row):
            continue
        fahrer_id = _clean_text(row[col_id])
        fahrer_name = _clean_text(row[col_name])
        if not fahrer_id or not fahrer_name or not fahrer_id.upper().startswith("F"):
            continue

        company_name = _clean_text(row[col_company]) if col_company is not None and col_company < len(row) else None
        status_entlassen = _clean_text(row[col_status]) if col_status is not None and col_status < len(row) else None
        datum_entlassen = _parse_date(row[col_dismiss]) if col_dismiss is not None and col_dismiss < len(row) else None

        for zero_based_col, report_year, iso_week in week_columns:
            week_start = date.fromisocalendar(report_year, iso_week, 1)
            week_end = date.fromisocalendar(report_year, iso_week, 7)
            week_code = _normalize_fahrer_week_code(row[zero_based_col] if zero_based_col < len(row) else None)
            is_active_in_week = datum_entlassen is None or datum_entlassen > week_start
            rows.append(
                FahrerWeekStatusRow(
                    report_year=report_year,
                    iso_week=iso_week,
                    week_start=week_start,
                    week_end=week_end,
                    fahrer_id=fahrer_id,
                    fahrer_name=fahrer_name,
                    company_name=company_name,
                    status_entlassen=status_entlassen,
                    datum_entlassen=datum_entlassen,
                    week_code=week_code,
                    is_active_in_week=is_active_in_week,
                    raw_payload={
                        "sheet": "Fahrer",
                        "row": row_idx,
                        "column": zero_based_col + 1,
                    },
                )
            )

    return sorted(rows, key=lambda r: (r.report_year, r.iso_week, r.fahrer_id.upper()))


def extract_einnahmen_months(wb) -> list[EinnahmenMonthRow]:
    if BERICHT_DISPO_SHEET not in wb.sheetnames:
        return []

    ws = wb[BERICHT_DISPO_SHEET]
    month_columns = _discover_month_columns(ws, preferred_row=6)
    if len(month_columns) < 2:
        return []

    nahverkehr_row = _find_metric_row(ws, ("Nahverkehr",))
    logistics_row = _find_metric_row(ws, ("Logistics",))
    gesamt_row = _find_metric_row(ws, ("Gesamt", "Total"))
    if nahverkehr_row is None or logistics_row is None or gesamt_row is None:
        return []

    rows: list[EinnahmenMonthRow] = []
    for month_idx in sorted(month_columns.keys()):
        month_name, col_idx = month_columns[month_idx]
        nah = _parse_decimal(ws.cell(row=nahverkehr_row, column=col_idx).value)
        log = _parse_decimal(ws.cell(row=logistics_row, column=col_idx).value)
        ges = _parse_decimal(ws.cell(row=gesamt_row, column=col_idx).value)

        rows.append(
            EinnahmenMonthRow(
                month_index=month_idx,
                month_name=month_name,
                nahverkehr=nah,
                logistics=log,
                gesamt=ges,
                raw_payload={
                    "sheet": BERICHT_DISPO_SHEET,
                    "month_column": col_idx,
                    "nahverkehr_row": nahverkehr_row,
                    "logistics_row": logistics_row,
                    "gesamt_row": gesamt_row,
                },
            )
        )
    return rows


def extract_einnahmen_firm_rows(wb) -> list[EinnahmenFirmRow]:
    if BERICHT_DISPO_SHEET not in wb.sheetnames:
        return []

    ws = wb[BERICHT_DISPO_SHEET]
    col_map = {
        "firm_name": 71,   # BS
        "january": 72,     # BT
        "february": 73,    # BU
        "march": 74,       # BV
        "april": 75,       # BW
        "may": 76,         # BX
        "june": 77,        # BY
        "july": 78,        # BZ
        "august": 79,      # CA
        "september": 80,   # CB
        "october": 81,     # CC
        "november": 82,    # CD
        "december": 83,    # CE
        "total": 84,       # CF
    }

    rows: list[EinnahmenFirmRow] = []
    for row_idx in range(3, 23):  # first 20 rows after header
        firm_name = _clean_text(ws.cell(row=row_idx, column=col_map["firm_name"]).value)
        if not firm_name:
            continue

        raw_payload = {
            "sheet": BERICHT_DISPO_SHEET,
            "source_range": f"BS{row_idx}:CF{row_idx}",
        }
        for key, col_idx in col_map.items():
            raw_payload[key] = ws.cell(row=row_idx, column=col_idx).value

        rows.append(
            EinnahmenFirmRow(
                row_index=row_idx - 2,
                firm_name=firm_name,
                january=_parse_decimal(ws.cell(row=row_idx, column=col_map["january"]).value),
                february=_parse_decimal(ws.cell(row=row_idx, column=col_map["february"]).value),
                march=_parse_decimal(ws.cell(row=row_idx, column=col_map["march"]).value),
                april=_parse_decimal(ws.cell(row=row_idx, column=col_map["april"]).value),
                may=_parse_decimal(ws.cell(row=row_idx, column=col_map["may"]).value),
                june=_parse_decimal(ws.cell(row=row_idx, column=col_map["june"]).value),
                july=_parse_decimal(ws.cell(row=row_idx, column=col_map["july"]).value),
                august=_parse_decimal(ws.cell(row=row_idx, column=col_map["august"]).value),
                september=_parse_decimal(ws.cell(row=row_idx, column=col_map["september"]).value),
                october=_parse_decimal(ws.cell(row=row_idx, column=col_map["october"]).value),
                november=_parse_decimal(ws.cell(row=row_idx, column=col_map["november"]).value),
                december=_parse_decimal(ws.cell(row=row_idx, column=col_map["december"]).value),
                total=_parse_decimal(ws.cell(row=row_idx, column=col_map["total"]).value),
                raw_payload=raw_payload,
            )
        )

    return rows


def _bonus_header_token(value: object) -> str:
    return str(value or "").strip().upper().replace("\u00a0", "").replace(" ", "").replace("％", "%")


def _parse_month_start(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)

    raw = str(value).strip()
    if not raw:
        return None

    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            d = datetime.strptime(raw, fmt).date()
            return date(d.year, d.month, 1)
        except ValueError:
            pass

    m = re.match(r"^\s*([A-Za-zА-Яа-яÄÖÜäöü]+)\s*[-./ ]\s*(\d{2,4})\s*$", raw)
    if m:
        month_token, year_token = m.group(1), m.group(2)
        parsed_month = _parse_month_token(month_token)
        if parsed_month:
            month_idx = parsed_month[0]
            year_num = int(year_token)
            if year_num < 100:
                year_num += 2000
            if 1970 <= year_num <= 2100:
                return date(year_num, month_idx, 1)

    return None


def _parse_percent_cell(ws, row_idx: int, col_idx: int) -> Decimal:
    cell = ws.cell(row=row_idx, column=col_idx)
    parsed = _parse_decimal(cell.value)
    if "%" in str(cell.number_format or "") and abs(parsed) <= Decimal("3"):
        parsed = (parsed * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return parsed


def _parse_int_cell(ws, row_idx: int, col_idx: int) -> int:
    parsed = _parse_decimal(ws.cell(row=row_idx, column=col_idx).value)
    return int(parsed.to_integral_value(rounding=ROUND_HALF_UP))


def extract_bonus_dynamik_months(wb) -> list[BonusDynamikRow]:
    if BONUS_DYNAMIK_SHEET not in wb.sheetnames:
        return []

    ws = wb[BONUS_DYNAMIK_SHEET]
    max_col = int(ws.max_column or 0)
    max_row = int(ws.max_row or 0)
    if max_col < 10 or max_row < 3:
        return []

    expected_headers = ["DAYS", "KM", "%KM", "CT", "%CT", "BONUS", "PENALTY", "FINAL"]
    month_blocks: list[tuple[int, date]] = []
    seen_months: set[tuple[int, int]] = set()
    for start_col in range(3, max_col - 7 + 1):
        sequence = [_bonus_header_token(ws.cell(row=2, column=start_col + off).value) for off in range(8)]
        if sequence != expected_headers:
            continue
        month_start = _parse_month_start(ws.cell(row=1, column=start_col).value)
        if not month_start:
            continue
        month_key = (month_start.year, month_start.month)
        if month_key in seen_months:
            continue
        seen_months.add(month_key)
        month_blocks.append((start_col, month_start))

    if not month_blocks:
        return []

    by_key: dict[tuple[int, int, str], BonusDynamikRow] = {}
    for row_idx in range(3, max_row + 1):
        fahrer_id = _clean_text(ws.cell(row=row_idx, column=1).value) or ""
        fahrer_name = _clean_text(ws.cell(row=row_idx, column=2).value) or ""
        if not fahrer_id and not fahrer_name:
            continue
        if not fahrer_id:
            continue

        for start_col, month_start in month_blocks:
            key = (month_start.year, month_start.month, fahrer_id.upper())
            by_key[key] = BonusDynamikRow(
                report_year=month_start.year,
                report_month=month_start.month,
                month_start=month_start,
                fahrer_id=fahrer_id,
                fahrer_name=fahrer_name,
                days=_parse_int_cell(ws, row_idx, start_col),
                km=_parse_decimal(ws.cell(row=row_idx, column=start_col + 1).value),
                pct_km=_parse_percent_cell(ws, row_idx, start_col + 2),
                ct=_parse_int_cell(ws, row_idx, start_col + 3),
                pct_ct=_parse_percent_cell(ws, row_idx, start_col + 4),
                bonus=_parse_decimal(ws.cell(row=row_idx, column=start_col + 5).value),
                penalty=_parse_decimal(ws.cell(row=row_idx, column=start_col + 6).value),
                final=_parse_decimal(ws.cell(row=row_idx, column=start_col + 7).value),
                raw_payload={
                    "sheet": BONUS_DYNAMIK_SHEET,
                    "row": row_idx,
                    "month_col": start_col,
                },
            )
    return [by_key[k] for k in sorted(by_key.keys())]


def _effective_header_values(ws, row_idx: int) -> list[str]:
    values: list[str] = []
    last = ""
    for col_idx in range(1, int(ws.max_column or 0) + 1):
        raw = _clean_text(ws.cell(row=row_idx, column=col_idx).value) or ""
        if raw:
            last = raw
        values.append(last)
    return values


def _find_first_diesel_column(top_headers: list[str], sub_headers: list[str], top_name: str, sub_name: str) -> int | None:
    target_top = _norm(top_name)
    target_sub = _norm(sub_name)
    for idx, (top, sub) in enumerate(zip(top_headers, sub_headers), start=1):
        if _norm(top) == target_top and _norm(sub) == target_sub:
            return idx
    return None


def _parse_int_like(value: object) -> int | None:
    parsed = _parse_decimal(value)
    out = int(parsed.to_integral_value(rounding=ROUND_HALF_UP))
    return out if out > 0 else None


def _parse_strict_positive_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, Decimal):
        if value != value.to_integral_value(rounding=ROUND_HALF_UP):
            return None
        out = int(value)
        return out if out > 0 else None
    if isinstance(value, float):
        if not value.is_integer():
            return None
        out = int(value)
        return out if out > 0 else None
    raw = str(value).strip()
    if not raw or not re.fullmatch(r"\d{1,4}", raw):
        return None
    out = int(raw)
    return out if out > 0 else None


def _parse_duration_minutes(value: object) -> int:
    if value is None:
        return 0

    if isinstance(value, timedelta):
        total_minutes = int((Decimal(str(value.total_seconds())) / Decimal("60")).to_integral_value(rounding=ROUND_HALF_UP))
        return total_minutes if total_minutes > 0 else 0

    if isinstance(value, datetime):
        return (value.hour * 60) + value.minute

    if isinstance(value, (int, float, Decimal)):
        parsed = Decimal(str(value))
        if parsed <= Decimal("0"):
            return 0
        # Excel durations are stored as day fractions.
        if abs(parsed) <= Decimal("10"):
            return int((parsed * Decimal("1440")).to_integral_value(rounding=ROUND_HALF_UP))
        return int(parsed.to_integral_value(rounding=ROUND_HALF_UP))

    raw = str(value).strip()
    if not raw or raw.upper() in {"#VALUE!", "#REF!", "#N/A", "N/A"}:
        return 0

    min_match = re.search(r"(-?\d+(?:[.,]\d+)?)\s*min", raw, flags=re.IGNORECASE)
    if min_match:
        return max(0, int(_parse_decimal(min_match.group(1)).to_integral_value(rounding=ROUND_HALF_UP)))

    hhmm_match = re.search(r"(-?\d+):(\d{1,2})", raw)
    if hhmm_match:
        hours = int(hhmm_match.group(1))
        minutes = int(hhmm_match.group(2))
        total = (hours * 60) + minutes
        return total if total > 0 else 0

    as_decimal = _parse_decimal(raw)
    if as_decimal > 0:
        if as_decimal <= Decimal("10"):
            return int((as_decimal * Decimal("1440")).to_integral_value(rounding=ROUND_HALF_UP))
        return int(as_decimal.to_integral_value(rounding=ROUND_HALF_UP))
    return 0


def extract_yf_fahrer_months(wb) -> list[YFFahrerMonthRow]:
    if YF_FAHRER_SHEET not in wb.sheetnames:
        return []

    ws = wb[YF_FAHRER_SHEET]
    header_row_idx = _find_header_row(ws, ("month", "fahrer"), max_scan_rows=10)
    header = _get_row_values(ws, header_row_idx)
    index = _build_col_index(header)

    col_month = _pick_col(index, "Month")
    col_fahrer = _pick_col(index, "Fahrer")
    col_distanz = _pick_col(index, "Distanz", "Distance")
    col_aktivitaet = _pick_col(index, "Aktivitätsdauer", "Aktivitatsdauer", "Aktivitaetsdauer")
    col_fahrzeit = _pick_col(index, "Fahrzeit")
    col_inaktiv = _pick_col(index, "Inaktivitätszeit", "Inaktivitatszeit", "Inaktivitaetszeit")

    required_cols = (col_month, col_fahrer, col_distanz, col_aktivitaet, col_fahrzeit, col_inaktiv)
    if any(col is None for col in required_cols):
        return []

    by_key: dict[tuple[int, str], YFFahrerMonthRow] = {}
    for row_idx, row in enumerate(_iter_sheet_rows(ws, header_row_idx), start=header_row_idx + 1):
        if col_month >= len(row) or col_fahrer >= len(row):
            continue
        month_index = _parse_int_like(row[col_month])
        fahrer_name = _clean_text(row[col_fahrer]) or ""
        if not month_index or month_index < 1 or month_index > 12 or not fahrer_name:
            continue

        payload = {
            str(header[i]).strip() if i < len(header) and header[i] is not None else f"col_{i+1}":
            (row[i] if i < len(row) else None)
            for i in range(len(header))
        }
        payload["sheet"] = YF_FAHRER_SHEET
        payload["row"] = row_idx

        key = (month_index, fahrer_name.upper())
        distanz_km = _parse_decimal(row[col_distanz]) if col_distanz < len(row) else Decimal("0.00")
        aktivitaet_total_minutes = _parse_duration_minutes(row[col_aktivitaet]) if col_aktivitaet < len(row) else 0
        fahrzeit_total_minutes = _parse_duration_minutes(row[col_fahrzeit]) if col_fahrzeit < len(row) else 0
        inaktivitaet_total_minutes = _parse_duration_minutes(row[col_inaktiv]) if col_inaktiv < len(row) else 0

        existing = by_key.get(key)
        if existing is None:
            by_key[key] = YFFahrerMonthRow(
                month_index=month_index,
                fahrer_name=fahrer_name,
                distanz_km=distanz_km,
                aktivitaet_total_minutes=aktivitaet_total_minutes,
                fahrzeit_total_minutes=fahrzeit_total_minutes,
                inaktivitaet_total_minutes=inaktivitaet_total_minutes,
                raw_payload=payload,
            )
        else:
            existing.distanz_km += distanz_km
            existing.aktivitaet_total_minutes += aktivitaet_total_minutes
            existing.fahrzeit_total_minutes += fahrzeit_total_minutes
            existing.inaktivitaet_total_minutes += inaktivitaet_total_minutes
            existing.raw_payload = payload

    return [by_key[k] for k in sorted(by_key.keys())]


def extract_yf_lkw_days(wb) -> list[YFLkwDayRow]:
    if YF_SHEET not in wb.sheetnames:
        return []

    ws = wb[YF_SHEET]
    header_row_idx = _find_header_row(ws, ("year", "lkw", "month", "datum"), max_scan_rows=10)
    header = _get_row_values(ws, header_row_idx)
    index = _build_col_index(header)

    col_year = _pick_col(index, "Year")
    col_lkw = _pick_col(index, "LKW")
    col_month = _pick_col(index, "Month")
    col_datum = _pick_col(index, "Datum", "Date")
    col_week = _pick_col(index, "Week")
    col_dayweek = _pick_col(index, "dayweek", "Dayweek", "Day of week")
    col_strecke = _pick_col(index, "Strecke", "Distance")
    col_km_start = _pick_col(index, "Kilometerstand Start")
    col_km_end = _pick_col(index, "Kilometerstand Ende")
    col_drivers_final = _pick_col(index, "Drivers final")

    required_cols = (
        col_year, col_lkw, col_month, col_datum, col_week,
        col_dayweek, col_strecke, col_km_start, col_km_end, col_drivers_final,
    )
    if any(col is None for col in required_cols):
        return []

    rows: list[YFLkwDayRow] = []
    for row_idx, row in enumerate(_iter_sheet_rows(ws, header_row_idx), start=header_row_idx + 1):
        if col_year >= len(row) or col_lkw >= len(row) or col_datum >= len(row):
            continue
        report_year = _parse_int_like(row[col_year])
        month_index = _parse_int_like(row[col_month]) if col_month < len(row) else None
        iso_week = _parse_int_like(row[col_week]) if col_week < len(row) else None
        lkw_nummer = _clean_text(row[col_lkw]) or ""
        report_date = _parse_date(row[col_datum])
        if not report_year or not month_index or not iso_week or not lkw_nummer or report_date is None:
            continue

        payload = {
            str(header[i]).strip() if i < len(header) and header[i] is not None else f"col_{i+1}":
            (row[i] if i < len(row) else None)
            for i in range(len(header))
        }
        payload["sheet"] = YF_SHEET
        payload["row"] = row_idx

        rows.append(
            YFLkwDayRow(
                report_year=report_year,
                month_index=month_index,
                month_name=MONTH_NAMES_DE[month_index - 1] if 1 <= month_index <= 12 else str(month_index),
                iso_week=iso_week,
                lkw_nummer=lkw_nummer,
                report_date=report_date,
                source_row=row_idx,
                dayweek=_clean_text(row[col_dayweek]) if col_dayweek < len(row) else None,
                strecke_km=_parse_decimal(row[col_strecke]) if col_strecke < len(row) else Decimal("0.00"),
                km_start=_parse_decimal(row[col_km_start]) if col_km_start < len(row) else Decimal("0.00"),
                km_end=_parse_decimal(row[col_km_end]) if col_km_end < len(row) else Decimal("0.00"),
                drivers_final=_clean_text(row[col_drivers_final]) if col_drivers_final < len(row) else None,
                raw_payload=payload,
            )
        )

    return sorted(rows, key=lambda r: (r.report_year, r.iso_week, r.lkw_nummer.upper(), r.report_date))


def extract_diesel_months(wb) -> list[DieselMonthRow]:
    if DIESEL_SHEET not in wb.sheetnames:
        return []

    ws = wb[DIESEL_SHEET]
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    if max_row < 3 or max_col < 17:
        return []

    top_headers = _effective_header_values(ws, 1)
    sub_headers = [str(ws.cell(row=2, column=col_idx).value or "").strip() for col_idx in range(1, max_col + 1)]

    col_month = _find_first_diesel_column(top_headers, sub_headers, "Diesel", "Month")
    col_year = _find_first_diesel_column(top_headers, sub_headers, "Diesel", "Year")
    col_liter_staack = _find_first_diesel_column(top_headers, sub_headers, "Liter", "Staack")
    col_liter_shell = _find_first_diesel_column(top_headers, sub_headers, "Liter", "Shell")
    col_liter_dkv = _find_first_diesel_column(top_headers, sub_headers, "Liter", "DKV")
    col_liter_total = _find_first_diesel_column(top_headers, sub_headers, "Liter", "Total")
    col_euro_staack = _find_first_diesel_column(top_headers, sub_headers, "Euro", "Staack")
    col_euro_shell = _find_first_diesel_column(top_headers, sub_headers, "Euro", "Shell")
    col_euro_dkv = _find_first_diesel_column(top_headers, sub_headers, "Euro", "DKV")
    col_euro_total = _find_first_diesel_column(top_headers, sub_headers, "Euro", "Total")
    col_eurpl_staack = _find_first_diesel_column(top_headers, sub_headers, "Euro/Liter", "Staack")
    col_eurpl_shell = _find_first_diesel_column(top_headers, sub_headers, "Euro/Liter", "Shell")
    col_eurpl_dkv = _find_first_diesel_column(top_headers, sub_headers, "Euro/Liter", "DKV")
    col_eurpl_avg = _find_first_diesel_column(top_headers, sub_headers, "Euro/Liter", "Average")

    required_cols = (
        col_month,
        col_year,
        col_liter_staack,
        col_liter_shell,
        col_liter_dkv,
        col_liter_total,
        col_euro_staack,
        col_euro_shell,
        col_euro_dkv,
        col_euro_total,
        col_eurpl_staack,
        col_eurpl_shell,
        col_eurpl_dkv,
        col_eurpl_avg,
    )
    if any(col is None for col in required_cols):
        return []

    rows: list[DieselMonthRow] = []
    for row_idx in range(3, max_row + 1):
        month_index = _parse_int_like(ws.cell(row=row_idx, column=col_month).value)
        report_year = _parse_int_like(ws.cell(row=row_idx, column=col_year).value)
        if not month_index or not report_year or month_index < 1 or month_index > 12:
            continue

        raw_payload = {"sheet": DIESEL_SHEET, "row": row_idx}
        for col_idx in range(1, max_col + 1):
            top_label = top_headers[col_idx - 1] or "col"
            sub_label = sub_headers[col_idx - 1] or str(col_idx)
            raw_payload[f"c{col_idx}:{top_label}::{sub_label}"] = ws.cell(row=row_idx, column=col_idx).value

        rows.append(
            DieselMonthRow(
                report_year=report_year,
                month_index=month_index,
                month_name=MONTH_NAMES_DE[month_index - 1],
                liter_staack=_parse_decimal(ws.cell(row=row_idx, column=col_liter_staack).value),
                liter_shell=_parse_decimal(ws.cell(row=row_idx, column=col_liter_shell).value),
                liter_dkv=_parse_decimal(ws.cell(row=row_idx, column=col_liter_dkv).value),
                liter_total=_parse_decimal(ws.cell(row=row_idx, column=col_liter_total).value),
                euro_staack=_parse_decimal(ws.cell(row=row_idx, column=col_euro_staack).value),
                euro_shell=_parse_decimal(ws.cell(row=row_idx, column=col_euro_shell).value),
                euro_dkv=_parse_decimal(ws.cell(row=row_idx, column=col_euro_dkv).value),
                euro_total=_parse_decimal(ws.cell(row=row_idx, column=col_euro_total).value),
                euro_per_liter_staack=_parse_decimal(ws.cell(row=row_idx, column=col_eurpl_staack).value),
                euro_per_liter_shell=_parse_decimal(ws.cell(row=row_idx, column=col_eurpl_shell).value),
                euro_per_liter_dkv=_parse_decimal(ws.cell(row=row_idx, column=col_eurpl_dkv).value),
                euro_per_liter_avg=_parse_decimal(ws.cell(row=row_idx, column=col_eurpl_avg).value),
                raw_payload=raw_payload,
            )
        )

    return sorted(rows, key=lambda r: (r.report_year, r.month_index))


def _row_value(row: list[object], zero_based_idx: int) -> object:
    return row[zero_based_idx] if 0 <= zero_based_idx < len(row) else None


def _row_payload(header: list[object], row: list[object], sheet: str, source_row: int) -> dict[str, object]:
    payload = {
        str(header[i]).strip() if i < len(header) and header[i] is not None else f"col_{i+1}":
        row[i] if i < len(row) else None
        for i in range(len(header))
    }
    payload["sheet"] = sheet
    payload["source_row"] = source_row
    return payload


def extract_lkw_fuel_transactions(wb) -> list[LkwFuelTransactionRow]:
    rows: list[LkwFuelTransactionRow] = []

    if STAACK_SHEET in wb.sheetnames:
        ws = wb[STAACK_SHEET]
        header = _get_row_values(ws, 1)
        # User-provided Staack structure, converted to zero-based indexes.
        idx_year, idx_month, idx_week = 0, 1, 2
        idx_product, idx_date, idx_quantity, idx_total_net = 18, 20, 22, 28
        idx_lkw, idx_driver = 43, 46
        for row_idx, row in enumerate(_iter_sheet_rows(ws, 1), start=2):
            product = _normalize_fuel_product(_row_value(row, idx_product))
            lkw_number = _normalize_lkw_number(_row_value(row, idx_lkw))
            if not product or not lkw_number:
                continue
            report_year = _parse_int_like(_row_value(row, idx_year))
            report_month = _parse_int_like(_row_value(row, idx_month))
            iso_week = _parse_int_like(_row_value(row, idx_week)) or 1
            transaction_date = _parse_date(_row_value(row, idx_date))
            if transaction_date:
                report_year = report_year or transaction_date.year
                report_month = report_month or transaction_date.month
                iso_week = iso_week or int(transaction_date.isocalendar()[1])
            if not report_year or not report_month or not (1 <= report_month <= 12):
                continue
            rows.append(
                LkwFuelTransactionRow(
                    source=STAACK_SHEET,
                    source_row=row_idx,
                    report_year=report_year,
                    report_month=report_month,
                    iso_week=iso_week if 1 <= iso_week <= 53 else 1,
                    transaction_date=transaction_date,
                    lkw_number=lkw_number,
                    product_name=product,
                    quantity_liters=_parse_decimal(_row_value(row, idx_quantity)),
                    total_net=_parse_decimal(_row_value(row, idx_total_net)),
                    driver_name=_clean_text(_row_value(row, idx_driver)),
                    raw_payload=_row_payload(header, row, STAACK_SHEET, row_idx),
                )
            )

    if SHELL_SHEET in wb.sheetnames:
        ws = wb[SHELL_SHEET]
        header = _get_row_values(ws, 1)
        index = _build_col_index(header)
        idx_year = _pick_col(index, "Year")
        idx_month = _pick_col(index, "Month")
        idx_week = _pick_col(index, "Week")
        idx_date = _pick_col(index, "Lieferdatum", "Date")
        idx_lkw = _pick_col(index, "KFZ-Kennzeichen", "CardLicanse Tag", "CardLicense Tag")
        idx_quantity = _pick_col(index, "Menge", "Quantity")
        idx_total_net = _pick_col(index, "NettobetraginTransaktionswährung", "Nettobetrag in Transaktionswährung")
        idx_product = _pick_col(index, "Produktname", "Product Name")
        idx_driver = _pick_col(index, "Fahrername", "Driver")
        required = (idx_year, idx_month, idx_week, idx_date, idx_lkw, idx_quantity, idx_total_net, idx_product)
        if all(col is not None for col in required):
            for row_idx, row in enumerate(_iter_sheet_rows(ws, 1), start=2):
                product = _normalize_fuel_product(_row_value(row, idx_product))
                lkw_number = _normalize_lkw_number(_row_value(row, idx_lkw))
                if not product or not lkw_number:
                    continue
                report_year = _parse_int_like(_row_value(row, idx_year))
                report_month = _parse_int_like(_row_value(row, idx_month))
                iso_week = _parse_int_like(_row_value(row, idx_week)) or 1
                transaction_date = _parse_date(_row_value(row, idx_date))
                if transaction_date:
                    report_year = report_year or transaction_date.year
                    report_month = report_month or transaction_date.month
                    iso_week = iso_week or int(transaction_date.isocalendar()[1])
                if not report_year or not report_month or not (1 <= report_month <= 12):
                    continue
                rows.append(
                    LkwFuelTransactionRow(
                        source=SHELL_SHEET,
                        source_row=row_idx,
                        report_year=report_year,
                        report_month=report_month,
                        iso_week=iso_week if 1 <= iso_week <= 53 else 1,
                        transaction_date=transaction_date,
                        lkw_number=lkw_number,
                        product_name=product,
                        quantity_liters=_parse_decimal(_row_value(row, idx_quantity)),
                        total_net=_parse_decimal(_row_value(row, idx_total_net)),
                        driver_name=_clean_text(_row_value(row, idx_driver)) if idx_driver is not None else None,
                        raw_payload=_row_payload(header, row, SHELL_SHEET, row_idx),
                    )
                )

    return sorted(rows, key=lambda r: (r.report_year, r.report_month, r.source, r.lkw_number, r.source_row))


def extract_lkw_revenue_rows(wb) -> list[LkwRevenueRow]:
    rows: list[LkwRevenueRow] = []

    if CARLO_SHEET in wb.sheetnames:
        ws = wb[CARLO_SHEET]
        header = _get_row_values(ws, 1)
        index = _build_col_index(header)
        idx_year = _pick_col(index, "Year")
        idx_month = _pick_col(index, "Month")
        idx_week = _pick_col(index, "Week")
        idx_lkw = _pick_col(index, "LKW(Soll)", "LKW Soll", "LKW")
        idx_amount = _pick_col(index, "Rechnung Betrag", "RechnungBetrag")
        if all(col is not None for col in (idx_year, idx_month, idx_week, idx_lkw, idx_amount)):
            for row_idx, row in enumerate(_iter_sheet_rows(ws, 1), start=2):
                lkw_number = _normalize_lkw_number(_row_value(row, idx_lkw))
                if not lkw_number:
                    continue
                report_year = _parse_int_like(_row_value(row, idx_year))
                report_month = _parse_int_like(_row_value(row, idx_month))
                iso_week = _parse_int_like(_row_value(row, idx_week)) or 1
                if not report_year or not report_month or not (1 <= report_month <= 12):
                    continue
                rows.append(
                    LkwRevenueRow(
                        source=CARLO_SHEET,
                        source_row=row_idx,
                        report_year=report_year,
                        report_month=report_month,
                        iso_week=iso_week if 1 <= iso_week <= 53 else 1,
                        lkw_number=lkw_number,
                        revenue_amount=_parse_decimal(_row_value(row, idx_amount)),
                        raw_payload=_row_payload(header, row, CARLO_SHEET, row_idx),
                    )
                )

    if CONTADO_SHEET in wb.sheetnames:
        ws = wb[CONTADO_SHEET]
        header = _get_row_values(ws, 1)
        index = _build_col_index(header)
        idx_lkw = _pick_col(index, "LKW")
        idx_year = _pick_col(index, "Year")
        idx_month = _pick_col(index, "Month")
        idx_week = _pick_col(index, "Week")
        idx_amount = _pick_col(index, "Kosten")
        if all(col is not None for col in (idx_lkw, idx_year, idx_month, idx_week, idx_amount)):
            for row_idx, row in enumerate(_iter_sheet_rows(ws, 1), start=2):
                lkw_number = _normalize_lkw_number(_row_value(row, idx_lkw))
                if not lkw_number:
                    continue
                report_year = _parse_int_like(_row_value(row, idx_year))
                report_month = _parse_int_like(_row_value(row, idx_month))
                iso_week = _parse_int_like(_row_value(row, idx_week)) or 1
                if not report_year or not report_month or not (1 <= report_month <= 12):
                    continue
                rows.append(
                    LkwRevenueRow(
                        source=CONTADO_SHEET,
                        source_row=row_idx,
                        report_year=report_year,
                        report_month=report_month,
                        iso_week=iso_week if 1 <= iso_week <= 53 else 1,
                        lkw_number=lkw_number,
                        revenue_amount=_parse_decimal(_row_value(row, idx_amount)),
                        raw_payload=_row_payload(header, row, CONTADO_SHEET, row_idx),
                    )
                )

    return sorted(rows, key=lambda r: (r.report_year, r.report_month, r.source, r.lkw_number, r.source_row))


def _prepare_readable_xlsm(source_path: Path) -> tuple[Path, bool]:
    """
    Returns (path, is_temp_copy_created_by_this_run).
    """
    if not source_path.exists():
        raise FileNotFoundError(f"XLSM source file not found: {source_path}")

    temp_dir = Path(tempfile.gettempdir())
    run_copy = temp_dir / f"LKW_Fahrer_Data_ETL_{int(time.time())}.xlsm"
    try:
        shutil.copy2(source_path, run_copy)
        return run_copy, True
    except Exception:
        pass

    fallback_candidates = sorted(
        glob.glob(str(temp_dir / "LKW_Fahrer_Data*.xlsm")),
        key=lambda p: Path(p).stat().st_mtime,
        reverse=True,
    )
    if fallback_candidates:
        return Path(fallback_candidates[0]), False

    raise PermissionError(
        "Could not create temp copy from EXCEL_FILE_PATH and no fallback copy was found in %TEMP%."
    )


def _upsert_company(cur, name: str) -> int:
    cur.execute(
        """
        INSERT INTO companies (name)
        VALUES (%s)
        ON CONFLICT (name) DO UPDATE SET updated_at = NOW()
        RETURNING id
        """,
        (name,),
    )
    return int(cur.fetchone()[0])


def _ensure_einnahmen_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_einnahmen_monthly (
            month_index SMALLINT PRIMARY KEY CHECK (month_index BETWEEN 1 AND 12),
            month_name TEXT NOT NULL,
            nahverkehr NUMERIC(14, 2) NOT NULL DEFAULT 0,
            logistics NUMERIC(14, 2) NOT NULL DEFAULT 0,
            gesamt NUMERIC(14, 2) NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """
    )


def _ensure_einnahmen_firm_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_einnahmen_firm_monthly (
            row_index SMALLINT PRIMARY KEY CHECK (row_index BETWEEN 1 AND 20),
            firm_name TEXT NOT NULL,
            january NUMERIC(14, 2) NOT NULL DEFAULT 0,
            february NUMERIC(14, 2) NOT NULL DEFAULT 0,
            march NUMERIC(14, 2) NOT NULL DEFAULT 0,
            april NUMERIC(14, 2) NOT NULL DEFAULT 0,
            may NUMERIC(14, 2) NOT NULL DEFAULT 0,
            june NUMERIC(14, 2) NOT NULL DEFAULT 0,
            july NUMERIC(14, 2) NOT NULL DEFAULT 0,
            august NUMERIC(14, 2) NOT NULL DEFAULT 0,
            september NUMERIC(14, 2) NOT NULL DEFAULT 0,
            october NUMERIC(14, 2) NOT NULL DEFAULT 0,
            november NUMERIC(14, 2) NOT NULL DEFAULT 0,
            december NUMERIC(14, 2) NOT NULL DEFAULT 0,
            total NUMERIC(14, 2) NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """
    )


def _ensure_bonus_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_bonus_dynamik_monthly (
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            report_month SMALLINT NOT NULL CHECK (report_month BETWEEN 1 AND 12),
            month_start DATE NOT NULL,
            fahrer_id TEXT NOT NULL,
            fahrer_name TEXT NOT NULL,
            days INTEGER NOT NULL DEFAULT 0,
            km NUMERIC(14, 2) NOT NULL DEFAULT 0,
            pct_km NUMERIC(8, 2) NOT NULL DEFAULT 0,
            ct INTEGER NOT NULL DEFAULT 0,
            pct_ct NUMERIC(8, 2) NOT NULL DEFAULT 0,
            bonus NUMERIC(14, 2) NOT NULL DEFAULT 0,
            penalty NUMERIC(14, 2) NOT NULL DEFAULT 0,
            final NUMERIC(14, 2) NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (report_year, report_month, fahrer_id)
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_bonus_dynamik_lookup
            ON report_bonus_dynamik_monthly (report_year, report_month, fahrer_name)
        """
    )


def _ensure_diesel_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_diesel_monthly (
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            month_index SMALLINT NOT NULL CHECK (month_index BETWEEN 1 AND 12),
            month_name TEXT NOT NULL,
            liter_staack NUMERIC(14, 2) NOT NULL DEFAULT 0,
            liter_shell NUMERIC(14, 2) NOT NULL DEFAULT 0,
            liter_dkv NUMERIC(14, 2) NOT NULL DEFAULT 0,
            liter_total NUMERIC(14, 2) NOT NULL DEFAULT 0,
            euro_staack NUMERIC(14, 2) NOT NULL DEFAULT 0,
            euro_shell NUMERIC(14, 2) NOT NULL DEFAULT 0,
            euro_dkv NUMERIC(14, 2) NOT NULL DEFAULT 0,
            euro_total NUMERIC(14, 2) NOT NULL DEFAULT 0,
            euro_per_liter_staack NUMERIC(10, 4) NOT NULL DEFAULT 0,
            euro_per_liter_shell NUMERIC(10, 4) NOT NULL DEFAULT 0,
            euro_per_liter_dkv NUMERIC(10, 4) NOT NULL DEFAULT 0,
            euro_per_liter_avg NUMERIC(10, 4) NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (report_year, month_index)
        )
        """
    )


def _ensure_lkw_fuel_transactions_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_lkw_fuel_transactions (
            source TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            report_month SMALLINT NOT NULL CHECK (report_month BETWEEN 1 AND 12),
            iso_week SMALLINT NOT NULL CHECK (iso_week BETWEEN 1 AND 53),
            transaction_date DATE,
            lkw_number TEXT NOT NULL,
            product_name TEXT NOT NULL,
            quantity_liters NUMERIC(14, 2) NOT NULL DEFAULT 0,
            total_net NUMERIC(14, 2) NOT NULL DEFAULT 0,
            driver_name TEXT,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (source, source_row)
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_lkw_fuel_lookup
            ON report_lkw_fuel_transactions (lkw_number, report_year, report_month, product_name, source)
        """
    )


def _ensure_lkw_revenue_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_lkw_revenue_records (
            source TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            report_month SMALLINT NOT NULL CHECK (report_month BETWEEN 1 AND 12),
            iso_week SMALLINT NOT NULL CHECK (iso_week BETWEEN 1 AND 53),
            lkw_number TEXT NOT NULL,
            revenue_amount NUMERIC(14, 2) NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (source, source_row)
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_lkw_revenue_lookup
            ON report_lkw_revenue_records (lkw_number, report_year, report_month, source)
        """
    )


def _ensure_yf_fahrer_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_yf_fahrer_monthly (
            month_index SMALLINT NOT NULL CHECK (month_index BETWEEN 1 AND 12),
            fahrer_name TEXT NOT NULL,
            distanz_km NUMERIC(14, 2) NOT NULL DEFAULT 0,
            aktivitaet_total_minutes INTEGER NOT NULL DEFAULT 0,
            fahrzeit_total_minutes INTEGER NOT NULL DEFAULT 0,
            inaktivitaet_total_minutes INTEGER NOT NULL DEFAULT 0,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (month_index, fahrer_name)
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_yf_fahrer_lookup
            ON report_yf_fahrer_monthly (month_index, fahrer_name)
        """
    )


def _ensure_yf_lkw_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_yf_lkw_daily (
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            month_index SMALLINT NOT NULL CHECK (month_index BETWEEN 1 AND 12),
            month_name TEXT NOT NULL,
            iso_week SMALLINT NOT NULL CHECK (iso_week BETWEEN 1 AND 53),
            lkw_nummer TEXT NOT NULL,
            report_date DATE NOT NULL,
            source_row INTEGER NOT NULL DEFAULT 0,
            dayweek TEXT,
            strecke_km NUMERIC(14, 2) NOT NULL DEFAULT 0,
            km_start NUMERIC(14, 2) NOT NULL DEFAULT 0,
            km_end NUMERIC(14, 2) NOT NULL DEFAULT 0,
            drivers_final TEXT,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (report_year, iso_week, lkw_nummer, report_date, source_row)
        )
        """
    )
    cur.execute(
        """
        ALTER TABLE report_yf_lkw_daily
        ADD COLUMN IF NOT EXISTS source_row INTEGER NOT NULL DEFAULT 0
        """
    )
    cur.execute("ALTER TABLE report_yf_lkw_daily DROP CONSTRAINT IF EXISTS report_yf_lkw_daily_pkey")
    cur.execute(
        """
        ALTER TABLE report_yf_lkw_daily
        ADD PRIMARY KEY (report_year, iso_week, lkw_nummer, report_date, source_row)
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_yf_lkw_lookup
            ON report_yf_lkw_daily (report_year, iso_week, lkw_nummer, report_date, source_row)
        """
    )


def _ensure_fahrer_weekly_status_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_fahrer_weekly_status (
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            iso_week SMALLINT NOT NULL CHECK (iso_week BETWEEN 1 AND 53),
            week_start DATE NOT NULL,
            week_end DATE NOT NULL,
            fahrer_id TEXT NOT NULL,
            fahrer_name TEXT NOT NULL,
            company_name TEXT,
            status_entlassen TEXT,
            datum_entlassen DATE,
            week_code TEXT NOT NULL DEFAULT '',
            is_active_in_week BOOLEAN NOT NULL DEFAULT TRUE,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
            PRIMARY KEY (report_year, iso_week, fahrer_id)
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_fahrer_weekly_status_lookup
            ON report_fahrer_weekly_status (report_year, iso_week, week_code, is_active_in_week)
        """
    )


def _ensure_repair_table(cur) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS report_repair_records (
            source_row INTEGER PRIMARY KEY,
            report_year SMALLINT NOT NULL CHECK (report_year BETWEEN 2020 AND 2100),
            report_month SMALLINT NOT NULL CHECK (report_month BETWEEN 1 AND 12),
            iso_week SMALLINT NOT NULL CHECK (iso_week BETWEEN 1 AND 53),
            invoice_date DATE,
            truck_number TEXT NOT NULL,
            original_truck_number TEXT,
            repair_name TEXT,
            total_price NUMERIC(14, 2) NOT NULL DEFAULT 0,
            invoice TEXT,
            seller TEXT,
            buyer TEXT,
            kategorie TEXT,
            raw_payload JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
        )
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_repair_truck_date
            ON report_repair_records (truck_number, invoice_date, report_year, report_month, iso_week)
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_report_repair_total
            ON report_repair_records (total_price DESC)
        """
    )


def run_etl(database_url: str, xlsm_path: Path) -> dict[str, int]:
    psycopg = _lazy_import_psycopg()
    created_copy = False
    readable_path = xlsm_path
    log_id = None

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO etl_log (source_name, status, details)
                VALUES (%s, 'running', %s::jsonb)
                RETURNING id
                """,
                ("xlsm_lkw_fahrer_data", json.dumps({"source_path": str(xlsm_path)}, ensure_ascii=False)),
            )
            log_id = int(cur.fetchone()[0])
        conn.commit()

        try:
            readable_path, created_copy = _prepare_readable_xlsm(xlsm_path)
            wb = openpyxl.load_workbook(readable_path, read_only=True, data_only=True, keep_vba=False)
            trucks = extract_trucks(wb)
            drivers = extract_drivers(wb)
            fahrer_weekly_rows = extract_fahrer_weekly_statuses(wb)
            einnahmen_rows = extract_einnahmen_months(wb)
            einnahmen_firm_rows = extract_einnahmen_firm_rows(wb)
            bonus_rows = extract_bonus_dynamik_months(wb)
            diesel_rows = extract_diesel_months(wb)
            lkw_fuel_rows = extract_lkw_fuel_transactions(wb)
            lkw_revenue_rows = extract_lkw_revenue_rows(wb)
            yf_fahrer_rows = extract_yf_fahrer_months(wb)
            yf_lkw_rows = extract_yf_lkw_days(wb)
            repair_rows = extract_repairs(wb)
            wb.close()

            company_names = sorted(
                {
                    c.strip()
                    for c in [*(t.company_name for t in trucks), *(d.company_name for d in drivers)]
                    if c and c.strip()
                }
            )

            with conn.cursor() as cur:
                company_ids: dict[str, int] = {}
                for name in company_names:
                    company_ids[name] = _upsert_company(cur, name)

                _ensure_einnahmen_table(cur)
                _ensure_einnahmen_firm_table(cur)
                _ensure_bonus_table(cur)
                _ensure_diesel_table(cur)
                _ensure_lkw_fuel_transactions_table(cur)
                _ensure_lkw_revenue_table(cur)
                _ensure_yf_fahrer_table(cur)
                _ensure_yf_lkw_table(cur)
                _ensure_fahrer_weekly_status_table(cur)
                _ensure_repair_table(cur)

                rows_inserted = 0
                rows_updated = 0
                rows_deleted = 0

                for t in trucks:
                    company_id = company_ids.get(t.company_name or "")
                    cur.execute(
                        """
                        INSERT INTO trucks (
                            external_id, plate_number, truck_type, company_id, status, status_since,
                            is_active, source_row_hash, raw_payload, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, NULL, %s::jsonb, NOW())
                        ON CONFLICT (external_id) DO UPDATE SET
                            plate_number = EXCLUDED.plate_number,
                            truck_type = EXCLUDED.truck_type,
                            company_id = EXCLUDED.company_id,
                            status = EXCLUDED.status,
                            status_since = EXCLUDED.status_since,
                            is_active = EXCLUDED.is_active,
                            raw_payload = EXCLUDED.raw_payload,
                            updated_at = NOW()
                        RETURNING (xmax = 0) AS inserted
                        """,
                        (
                            t.external_id,
                            t.plate_number,
                            t.truck_type,
                            company_id,
                            t.status,
                            t.status_since,
                            t.is_active,
                            json.dumps(t.raw_payload, ensure_ascii=False),
                        ),
                    )
                    inserted = bool(cur.fetchone()[0])
                    rows_inserted += 1 if inserted else 0
                    rows_updated += 0 if inserted else 1

                for d in drivers:
                    company_id = company_ids.get(d.company_name or "")
                    cur.execute(
                        """
                        INSERT INTO drivers (
                            external_id, full_name, phone, company_id, is_active,
                            source_row_hash, raw_payload, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, NULL, %s::jsonb, NOW())
                        ON CONFLICT (external_id) DO UPDATE SET
                            full_name = EXCLUDED.full_name,
                            phone = EXCLUDED.phone,
                            company_id = EXCLUDED.company_id,
                            is_active = EXCLUDED.is_active,
                            raw_payload = EXCLUDED.raw_payload,
                            updated_at = NOW()
                        RETURNING (xmax = 0) AS inserted
                        """,
                        (
                            d.external_id,
                            d.full_name,
                            d.phone,
                            company_id,
                            d.is_active,
                            json.dumps(d.raw_payload, ensure_ascii=False),
                        ),
                    )
                    inserted = bool(cur.fetchone()[0])
                    rows_inserted += 1 if inserted else 0
                    rows_updated += 0 if inserted else 1

                cur.execute("DELETE FROM report_fahrer_weekly_status")
                rows_deleted += int(cur.rowcount or 0)
                for rec in fahrer_weekly_rows:
                    cur.execute(
                        """
                        INSERT INTO report_fahrer_weekly_status (
                            report_year,
                            iso_week,
                            week_start,
                            week_end,
                            fahrer_id,
                            fahrer_name,
                            company_name,
                            status_entlassen,
                            datum_entlassen,
                            week_code,
                            is_active_in_week,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.report_year,
                            rec.iso_week,
                            rec.week_start,
                            rec.week_end,
                            rec.fahrer_id,
                            rec.fahrer_name,
                            rec.company_name,
                            rec.status_entlassen,
                            rec.datum_entlassen,
                            rec.week_code,
                            rec.is_active_in_week,
                            json.dumps(rec.raw_payload, ensure_ascii=False, default=str),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_einnahmen_monthly")
                rows_deleted += int(cur.rowcount or 0)
                for rec in einnahmen_rows:
                    cur.execute(
                        """
                        INSERT INTO report_einnahmen_monthly (
                            month_index,
                            month_name,
                            nahverkehr,
                            logistics,
                            gesamt,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.month_index,
                            rec.month_name,
                            rec.nahverkehr,
                            rec.logistics,
                            rec.gesamt,
                            json.dumps(rec.raw_payload, ensure_ascii=False),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_einnahmen_firm_monthly")
                rows_deleted += int(cur.rowcount or 0)
                for rec in einnahmen_firm_rows:
                    cur.execute(
                        """
                        INSERT INTO report_einnahmen_firm_monthly (
                            row_index,
                            firm_name,
                            january,
                            february,
                            march,
                            april,
                            may,
                            june,
                            july,
                            august,
                            september,
                            october,
                            november,
                            december,
                            total,
                            raw_payload,
                            updated_at
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW()
                        )
                        """,
                        (
                            rec.row_index,
                            rec.firm_name,
                            rec.january,
                            rec.february,
                            rec.march,
                            rec.april,
                            rec.may,
                            rec.june,
                            rec.july,
                            rec.august,
                            rec.september,
                            rec.october,
                            rec.november,
                            rec.december,
                            rec.total,
                            json.dumps(rec.raw_payload, ensure_ascii=False),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_bonus_dynamik_monthly")
                rows_deleted += int(cur.rowcount or 0)
                for rec in bonus_rows:
                    cur.execute(
                        """
                        INSERT INTO report_bonus_dynamik_monthly (
                            report_year,
                            report_month,
                            month_start,
                            fahrer_id,
                            fahrer_name,
                            days,
                            km,
                            pct_km,
                            ct,
                            pct_ct,
                            bonus,
                            penalty,
                            final,
                            raw_payload,
                            updated_at
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW()
                        )
                        """,
                        (
                            rec.report_year,
                            rec.report_month,
                            rec.month_start,
                            rec.fahrer_id,
                            rec.fahrer_name,
                            rec.days,
                            rec.km,
                            rec.pct_km,
                            rec.ct,
                            rec.pct_ct,
                            rec.bonus,
                            rec.penalty,
                            rec.final,
                            json.dumps(rec.raw_payload, ensure_ascii=False),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_diesel_monthly")
                rows_deleted += int(cur.rowcount or 0)
                for rec in diesel_rows:
                    cur.execute(
                        """
                        INSERT INTO report_diesel_monthly (
                            report_year,
                            month_index,
                            month_name,
                            liter_staack,
                            liter_shell,
                            liter_dkv,
                            liter_total,
                            euro_staack,
                            euro_shell,
                            euro_dkv,
                            euro_total,
                            euro_per_liter_staack,
                            euro_per_liter_shell,
                            euro_per_liter_dkv,
                            euro_per_liter_avg,
                            raw_payload,
                            updated_at
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW()
                        )
                        """,
                        (
                            rec.report_year,
                            rec.month_index,
                            rec.month_name,
                            rec.liter_staack,
                            rec.liter_shell,
                            rec.liter_dkv,
                            rec.liter_total,
                            rec.euro_staack,
                            rec.euro_shell,
                            rec.euro_dkv,
                            rec.euro_total,
                            rec.euro_per_liter_staack,
                            rec.euro_per_liter_shell,
                            rec.euro_per_liter_dkv,
                            rec.euro_per_liter_avg,
                            json.dumps(rec.raw_payload, ensure_ascii=False),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_lkw_fuel_transactions")
                rows_deleted += int(cur.rowcount or 0)
                for rec in lkw_fuel_rows:
                    cur.execute(
                        """
                        INSERT INTO report_lkw_fuel_transactions (
                            source,
                            source_row,
                            report_year,
                            report_month,
                            iso_week,
                            transaction_date,
                            lkw_number,
                            product_name,
                            quantity_liters,
                            total_net,
                            driver_name,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.source,
                            rec.source_row,
                            rec.report_year,
                            rec.report_month,
                            rec.iso_week,
                            rec.transaction_date,
                            rec.lkw_number,
                            rec.product_name,
                            rec.quantity_liters,
                            rec.total_net,
                            rec.driver_name,
                            json.dumps(rec.raw_payload, ensure_ascii=False, default=str),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_lkw_revenue_records")
                rows_deleted += int(cur.rowcount or 0)
                for rec in lkw_revenue_rows:
                    cur.execute(
                        """
                        INSERT INTO report_lkw_revenue_records (
                            source,
                            source_row,
                            report_year,
                            report_month,
                            iso_week,
                            lkw_number,
                            revenue_amount,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.source,
                            rec.source_row,
                            rec.report_year,
                            rec.report_month,
                            rec.iso_week,
                            rec.lkw_number,
                            rec.revenue_amount,
                            json.dumps(rec.raw_payload, ensure_ascii=False, default=str),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_yf_fahrer_monthly")
                rows_deleted += int(cur.rowcount or 0)
                for rec in yf_fahrer_rows:
                    cur.execute(
                        """
                        INSERT INTO report_yf_fahrer_monthly (
                            month_index,
                            fahrer_name,
                            distanz_km,
                            aktivitaet_total_minutes,
                            fahrzeit_total_minutes,
                            inaktivitaet_total_minutes,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.month_index,
                            rec.fahrer_name,
                            rec.distanz_km,
                            rec.aktivitaet_total_minutes,
                            rec.fahrzeit_total_minutes,
                            rec.inaktivitaet_total_minutes,
                            json.dumps(rec.raw_payload, ensure_ascii=False, default=str),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_yf_lkw_daily")
                rows_deleted += int(cur.rowcount or 0)
                for rec in yf_lkw_rows:
                    cur.execute(
                        """
                        INSERT INTO report_yf_lkw_daily (
                            report_year,
                            month_index,
                            month_name,
                            iso_week,
                            lkw_nummer,
                            report_date,
                            source_row,
                            dayweek,
                            strecke_km,
                            km_start,
                            km_end,
                            drivers_final,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.report_year,
                            rec.month_index,
                            rec.month_name,
                            rec.iso_week,
                            rec.lkw_nummer,
                            rec.report_date,
                            rec.source_row,
                            rec.dayweek,
                            rec.strecke_km,
                            rec.km_start,
                            rec.km_end,
                            rec.drivers_final,
                            json.dumps(rec.raw_payload, ensure_ascii=False, default=str),
                        ),
                    )
                    rows_inserted += 1

                cur.execute("DELETE FROM report_repair_records")
                rows_deleted += int(cur.rowcount or 0)
                for rec in repair_rows:
                    cur.execute(
                        """
                        INSERT INTO report_repair_records (
                            source_row,
                            report_year,
                            report_month,
                            iso_week,
                            invoice_date,
                            truck_number,
                            original_truck_number,
                            repair_name,
                            total_price,
                            invoice,
                            seller,
                            buyer,
                            kategorie,
                            raw_payload,
                            updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, NOW())
                        """,
                        (
                            rec.source_row,
                            rec.report_year,
                            rec.report_month,
                            rec.iso_week,
                            rec.invoice_date,
                            rec.truck_number,
                            rec.original_truck_number,
                            rec.repair_name,
                            rec.total_price,
                            rec.invoice,
                            rec.seller,
                            rec.buyer,
                            rec.kategorie,
                            json.dumps(rec.raw_payload, ensure_ascii=False, default=str),
                        ),
                    )
                    rows_inserted += 1

                cur.execute(
                    """
                    UPDATE etl_log
                    SET
                        status = 'success',
                        finished_at = NOW(),
                        rows_read = %s,
                        rows_inserted = %s,
                        rows_updated = %s,
                        rows_deleted = %s,
                        details = %s::jsonb
                    WHERE id = %s
                    """,
                    (
                        len(trucks) + len(drivers) + len(fahrer_weekly_rows) + len(einnahmen_rows) + len(einnahmen_firm_rows) + len(bonus_rows) + len(diesel_rows) + len(lkw_fuel_rows) + len(lkw_revenue_rows) + len(yf_fahrer_rows) + len(yf_lkw_rows) + len(repair_rows),
                        rows_inserted,
                        rows_updated,
                        rows_deleted,
                        json.dumps(
                            {
                                "companies": len(company_names),
                                "trucks": len(trucks),
                                "drivers": len(drivers),
                                "fahrer_weekly_rows": len(fahrer_weekly_rows),
                                "einnahmen_months": len(einnahmen_rows),
                                "einnahmen_firms": len(einnahmen_firm_rows),
                                "bonus_rows": len(bonus_rows),
                                "diesel_months": len(diesel_rows),
                                "lkw_fuel_rows": len(lkw_fuel_rows),
                                "lkw_revenue_rows": len(lkw_revenue_rows),
                                "yf_fahrer_rows": len(yf_fahrer_rows),
                                "yf_lkw_rows": len(yf_lkw_rows),
                                "repair_rows": len(repair_rows),
                                "workbook_used": str(readable_path),
                            },
                            ensure_ascii=False,
                        ),
                        log_id,
                    ),
                )
            conn.commit()

            return {
                "companies": len(company_names),
                "trucks": len(trucks),
                "drivers": len(drivers),
                "fahrer_weekly_rows": len(fahrer_weekly_rows),
                "einnahmen_months": len(einnahmen_rows),
                "einnahmen_firms": len(einnahmen_firm_rows),
                "bonus_rows": len(bonus_rows),
                "diesel_months": len(diesel_rows),
                "lkw_fuel_rows": len(lkw_fuel_rows),
                "lkw_revenue_rows": len(lkw_revenue_rows),
                "yf_fahrer_rows": len(yf_fahrer_rows),
                "yf_lkw_rows": len(yf_lkw_rows),
                "repair_rows": len(repair_rows),
            }

        except Exception as exc:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE etl_log
                    SET status = 'failed', finished_at = NOW(), error_message = %s
                    WHERE id = %s
                    """,
                    (str(exc), log_id),
                )
            conn.commit()
            raise
        finally:
            if created_copy and readable_path.exists():
                try:
                    readable_path.unlink()
                except OSError:
                    pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Import LKW/Fahrer master data from XLSM to PostgreSQL.")
    parser.add_argument("--database-url", default="", help="Override DATABASE_URL from env")
    parser.add_argument("--xlsm-path", default="", help="Override EXCEL_FILE_PATH from env")
    args = parser.parse_args()

    load_dotenv(override=True)
    database_url = (args.database_url or os.getenv("DATABASE_URL", "")).strip()
    xlsm_raw = (args.xlsm_path or os.getenv("EXCEL_FILE_PATH", "")).strip()

    if not database_url:
        raise RuntimeError("DATABASE_URL is empty. Set it in .env or pass --database-url.")
    if not xlsm_raw:
        raise RuntimeError("EXCEL_FILE_PATH is empty. Set it in .env or pass --xlsm-path.")

    result = run_etl(database_url=database_url, xlsm_path=Path(xlsm_raw))
    print(
        f"ETL success: companies={result['companies']} "
        f"trucks={result['trucks']} drivers={result['drivers']} "
        f"fahrer_weekly_rows={result['fahrer_weekly_rows']} "
        f"einnahmen_months={result['einnahmen_months']} "
        f"einnahmen_firms={result['einnahmen_firms']} "
        f"bonus_rows={result['bonus_rows']} "
        f"diesel_months={result['diesel_months']} "
        f"lkw_fuel_rows={result['lkw_fuel_rows']} "
        f"lkw_revenue_rows={result['lkw_revenue_rows']} "
        f"yf_fahrer_rows={result['yf_fahrer_rows']} "
        f"yf_lkw_rows={result['yf_lkw_rows']} "
        f"repair_rows={result['repair_rows']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
