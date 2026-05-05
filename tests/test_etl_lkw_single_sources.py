from datetime import date
from decimal import Decimal

from openpyxl import Workbook

from etl_xlsm_to_postgres import extract_lkw_fuel_transactions, extract_lkw_revenue_rows


def test_extract_lkw_fuel_transactions_from_staack_and_shell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Staack"
    staack_headers = [None] * 63
    staack_headers[0] = "Year"
    staack_headers[1] = "Month"
    staack_headers[2] = "Week"
    staack_headers[18] = "Product\nName"
    staack_headers[20] = "Date"
    staack_headers[22] = "Quantity"
    staack_headers[28] = "Price\nTotal Net"
    staack_headers[43] = "CardLicense\nTag"
    staack_headers[46] = "Driver Card\nDriver"
    ws.append(staack_headers)
    staack_row = [None] * 63
    staack_row[0] = 2026
    staack_row[1] = 4
    staack_row[2] = 15
    staack_row[18] = "Diesel"
    staack_row[20] = date(2026, 4, 7)
    staack_row[22] = "310,13"
    staack_row[28] = "15,07"
    staack_row[43] = "GR-OO1708"
    staack_row[46] = "Driver A"
    ws.append(staack_row)

    shell = wb.create_sheet("Shell")
    shell_headers = [
        "Year",
        "Month",
        "Week",
        "x4",
        "x5",
        "x6",
        "x7",
        "Lieferdatum",
        "x9",
        "x10",
        "x11",
        "x12",
        "x13",
        "x14",
        "x15",
        "x16",
        "x17",
        "x18",
        "KFZ-Kennzeichen",
        "x20",
        "x21",
        "x22",
        "Menge",
        "x24",
        "x25",
        "x26",
        "x27",
        "x28",
        "x29",
        "x30",
        "x31",
        "x32",
        "x33",
        "x34",
        "NettobetraginTransaktionswährung",
        "x36",
        "x37",
        "x38",
        "x39",
        "x40",
        "x41",
        "x42",
        "x43",
        "x44",
        "x45",
        "x46",
        "x47",
        "x48",
        "x49",
        "x50",
        "x51",
        "x52",
        "x53",
        "x54",
        "x55",
        "x56",
        "x57",
        "x58",
        "x59",
        "x60",
        "Fahrername",
        "x62",
        "x63",
        "Produktname",
    ]
    shell.append(shell_headers)
    shell.append([2026, 4, 15, None, None, None, None, date(2026, 4, 8), None, None, None, None, None, None, None, None, None, None, "GR-OO1708", None, None, None, 42, None, None, None, None, None, None, None, None, None, None, None, 84.5, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Driver B", None, None, "AdBlue"])

    rows = extract_lkw_fuel_transactions(wb)

    assert len(rows) == 2
    assert {row.source for row in rows} == {"Staack", "Shell"}
    assert {row.product_name for row in rows} == {"Diesel", "AdBlue"}
    staack = next(row for row in rows if row.source == "Staack")
    assert staack.lkw_number == "GR-OO1708"
    assert staack.quantity_liters == Decimal("310.13")


def test_extract_lkw_revenue_rows_from_carlo_and_contado():
    wb = Workbook()
    carlo = wb.active
    carlo.title = "Carlo"
    carlo.append(["Year", "Month", "Week", "Auftragsnummer", "Transportart", "Ladedatum (Soll) Start", "Lieferdatum (Soll) Ende", "LKW(Soll)", "Rechnung Betrag"])
    carlo.append([2026, 4, 16, 9344, "PLANE", None, None, "GR-OO2236", "255,80"])

    contado = wb.create_sheet("Contado")
    contado.append(["LKW", "Year", "Week", "Month", "Empfänger", "Ort", "Debitornummer", "Kreditornummer", "Referenz", "L-Datum", "Containernummer", "KD-Ref", "Lauf-Ref", "Gr+Typ", "Bel/", "Aufnahme", "Gestellung", "Abgabe", "Tarifbasis", "Artikel", "Bemerkung", "USt", "Erlös", "Kosten"])
    contado.append(["GR-OO2236", 2026, 16, 4, "Groo Truck", None, None, None, None, None, None, None, None, None, None, None, None, None, None, "1 x Fracht", None, "RC", 0, 648.55])

    rows = extract_lkw_revenue_rows(wb)

    assert len(rows) == 2
    assert {row.source for row in rows} == {"Carlo", "Contado"}
    assert sum(row.revenue_amount for row in rows) == Decimal("904.35")
    assert all(row.lkw_number == "GR-OO2236" for row in rows)
