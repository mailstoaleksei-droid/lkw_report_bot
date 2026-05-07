from datetime import date

from openpyxl import Workbook

from etl_xlsm_to_postgres import extract_drivers, extract_fahrer_weekly_statuses


def _build_fahrer_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fahrer"

    ws["A1"] = ">>Neu Fahrer"
    ws["B1"] = ">>Search"
    ws["Z1"] = 2026
    ws["AH1"] = 2026

    headers = [
        "Fahrer-ID",
        "Fahrername",
        "Firma",
        "Telefonnummer",
        "Führerschein",
        "LKW-Typ",
        "Arbeitsplan",
        "Status",
        "Datum entlassen",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=value)

    ws["Z2"] = "Urlaub gesamt"
    ws["AA2"] = "Krankheitstage"
    ws["AH2"] = 2026
    ws["AI2"] = 2026
    ws["AJ2"] = 2026

    sub_headers = ["ID", "Name", "Company", "Phone", "License", "Type", "Schedule", "Active/Fired", "Date"]
    for idx, value in enumerate(sub_headers, start=1):
        ws.cell(row=3, column=idx, value=value)

    ws["Z3"] = "Total vacation"
    ws["AA3"] = "Sick Days"
    ws["AH3"] = 1
    ws["AI3"] = 2
    ws["AJ3"] = 3

    ws["A4"] = "F001"
    ws["B4"] = "Driver One"
    ws["C4"] = "Groo"
    ws["D4"] = "+491111"
    ws["F4"] = "Container"
    ws["G4"] = "3M/3M"
    ws["Z4"] = 10
    ws["AA4"] = 2
    ws["AH4"] = "U"
    ws["AI4"] = "U"
    ws["AJ4"] = "К"

    ws["A5"] = "F002"
    ws["B5"] = "Driver Two"
    ws["C5"] = "Groo"
    ws["D5"] = "+492222"
    ws["F5"] = "Planen"
    ws["G5"] = "2M/2M"
    ws["H5"] = "Fahrer entlassen"
    ws["I5"] = date(2026, 1, 12)
    ws["AH5"] = "U"
    ws["AI5"] = "K"
    ws["AJ5"] = "U"

    return wb


def _build_fahrer_sheet_with_current_card_dates():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fahrer"

    headers = [
        "Fahrer-ID",
        "Fahrername",
        "Firma",
        "Telefonnummer",
        "Führerschein",
        "LKW-Typ",
        "Arbeitsplan",
        "Eintrittsdatum\ndes Fahrers",
        "Status",
        "Datum entlassen",
        "Pass gültig bis",
        "95 Code\nrosa Papier bis",
        "Art der Wohnungen bis",
        "Fahrerkarte\ngültig bis",
    ]
    sub_headers = [
        "ID",
        "Name",
        "Company",
        "Phone",
        "License",
        "Type",
        "Schedule",
        "Driver start\ndate",
        "Active/Fired",
        "Date",
        "Passport valid",
        "95 code",
        "Type of residence",
        "Driver card valid until",
    ]
    for idx, value in enumerate(headers, start=1):
        ws.cell(row=2, column=idx, value=value)
    for idx, value in enumerate(sub_headers, start=1):
        ws.cell(row=3, column=idx, value=value)

    ws["A4"] = "F001"
    ws["B4"] = "Driver One"
    ws["C4"] = "Groo"
    ws["D4"] = "+491111"
    ws["F4"] = "Container"
    ws["G4"] = "3M/3M"
    ws["H4"] = date(2026, 1, 1)
    ws["N4"] = date(2033, 3, 4)

    return wb


def test_extract_fahrer_weekly_statuses_parses_all_driver_weeks():
    wb = _build_fahrer_sheet()

    rows = extract_fahrer_weekly_statuses(wb)

    assert len(rows) == 6
    first = rows[0]
    assert first.report_year == 2026
    assert first.iso_week == 1
    assert first.week_start == date.fromisocalendar(2026, 1, 1)
    assert first.week_end == date.fromisocalendar(2026, 1, 7)
    assert first.fahrer_id == "F001"
    assert first.week_code == "U"
    assert first.is_active_in_week is True


def test_extract_fahrer_weekly_statuses_normalizes_cyrillic_sick_code():
    wb = _build_fahrer_sheet()

    rows = extract_fahrer_weekly_statuses(wb)
    target = next(r for r in rows if r.fahrer_id == "F001" and r.iso_week == 3)

    assert target.week_code == "K"


def test_extract_fahrer_weekly_statuses_marks_dismissed_driver_inactive_from_week_start():
    wb = _build_fahrer_sheet()

    rows = extract_fahrer_weekly_statuses(wb)
    week2 = next(r for r in rows if r.fahrer_id == "F002" and r.iso_week == 2)
    week3 = next(r for r in rows if r.fahrer_id == "F002" and r.iso_week == 3)

    assert week2.week_start == date(2026, 1, 5)
    assert week2.is_active_in_week is True
    assert week3.week_start == date(2026, 1, 12)
    assert week3.is_active_in_week is False


def test_extract_drivers_preserves_current_fahrer_card_date_columns():
    wb = _build_fahrer_sheet_with_current_card_dates()

    rows = extract_drivers(wb)

    assert len(rows) == 1
    payload = rows[0].raw_payload
    assert payload["Eintrittsdatum\ndes Fahrers"] == "2026-01-01"
    assert payload["Fahrerkarte\ngültig bis"] == "2033-03-04"
